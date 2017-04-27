import os
import urllib2
import time
import re
from datetime import datetime
from bs4 import BeautifulSoup as bs
import openpyxl

def getLogs(logURL, txID):

	""" returns the logs file object """

	
	url = '%s' % str(logURL)
	page = urllib2.urlopen(url)

	# time.sleep(8)

	print logURL
	print txID

	print 'in get Logs'

	try:

		soup = bs(page.read(), 'html.parser')
		print len(soup)

		myLogs = []

		print '1'
		myLogs_firstLine = soup.find(text=re.compile('%s' % txID))
		print '2'
		myLogs_lastLine = soup.find_all(text=re.compile('%s' % txID))[-1]
		print '3'
		myLogs.append('%s\n' % myLogs_firstLine)

		print "parsing done"

	
		# print(len(soup.find_all('br')))

		for _ in range(10000):

			next_sib = myLogs_firstLine.next

			if next_sib != myLogs_lastLine:
				if next_sib.string != "None":
					myLogs.append("%s\n" % next_sib)
					myLogs_firstLine = next_sib
			else:
				break
		
		print 'no exception'
		myLogs.append(myLogs_lastLine)

	except Exception, e:
		print e
		return 1

	return myLogs


def writeToFile(foldername, ObjLog, dt, tm, txnName=None, filename=None):

	if txnName != None:
		directory = r'%s\\%s\\%s' % (foldername, txnName, dt, )
	else:
		directory = r'%s\\%s\\%s' % (foldername, dt, tm)

	print "writeToFile"
	print directory

	if not os.path.exists(directory):
		os.makedirs(directory)

	try:

		if txnName != None:
			filePath = r'%s\\%s_%s.txt' % (directory, tm, txnName, )

			f = open(filePath, 'a')

			# f.write('<pre>')
			for tex in ObjLog:
				if tex == '<br/>\n':
					continue

				elif tex[0:5] == "<font":
					continue

				else:
					f.write('%s' % tex)
			# f.write('</pre>')
			f.close()

		else:
			filePath = r'%s\\%s.txt' % (directory, filename, )

			f = open(filePath, 'a')

			for tex in ObjLog:
				f.write('%s' % tex)
			f.close()


		print "file created at %s" % filePath
		return filePath
	except Exception, e:
		print(e)
		return 1


def getDateTime():

	d = datetime.now().date().strftime("%Y.%m.%d")
	t = datetime.now().time().strftime("%H.%M.%S.%f")

	return [d, t]

def get_X12s(filename):
	""" Retrieves X12 requests """

	# Work Book from current work directory
	wb = openpyxl.load_workbook('%s' % filename)

	# Get the sheet - Requests
	sheet = wb.get_sheet_by_name("Requests")

	# List of all X12s
	X12_list = []

	# Maximum number of filled rows
	max_rows = sheet.max_row

	# Maximum number of filled columns
	max_cols = sheet.max_column

	# Get the X12s
	for i in range(2, max_rows + 1):
		
		X12_list.append(sheet.cell(row=i, column=2).value)		

	# Remove None from the List
	final_X12_list = [i for i in X12_list if i is not None]

	# Return the final list
	return final_X12_list


def get_TCs(filename):
	""" Retrieves TCs from the excel sheet """

	# Work Book from current work directory
	wb = openpyxl.load_workbook('%s' % filename)

	# Get the sheet - TestCases
	try:
		sheet = wb.get_sheet_by_name("TestCases")
	except Exception, e:
		print str(e)
		return [1, str(e)]

	# List of all TCss
	TCs_list = []

	# Maximum number of filled rows
	max_rows = sheet.max_row

	# Maximum number of filled columns
	max_cols = sheet.max_column

	# Get the TCs
	for i in range(2, max_rows + 1):

		temp=[]

		for j in range(2, max_cols + 1):
			
			temp.append(sheet.cell(row=i, column=j).value)

		TCs_list.append(temp)	

	# Remove None from the List
	final_TCs_list = [i for i in TCs_list if i is not None]

	print final_TCs_list

	# Return the final list
	return [0, final_TCs_list]